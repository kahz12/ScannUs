"""
utils/results_parse.py — Transforms search result sets into various output formats.

Formats supported:
  - Rich terminal table (to_table)
  - HTML report  (export_html)
  - CSV          (export_csv)
  - Excel .xlsx  (export_excel)  — styled: colors, autofit, freeze panes, hyperlinks
  - JSON         (export_json)
"""

import json
import os
import csv
import html
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.table import Table
from cli.ui import THEME, print_success, print_error, make_table
from core.config import DIR_REPORTS


_HTML_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>ScannUs — Search Report</title>
<style>
  :root { color-scheme: dark; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui, sans-serif;
         background: #0f1117; color: #e6e6e6; margin: 0; padding: 2rem; }
  header { border-bottom: 1px solid #2a2f3a; padding-bottom: 1rem; margin-bottom: 1.5rem; }
  header h1 { margin: 0 0 .25rem 0; font-size: 1.4rem; color: #00d7ff; letter-spacing: .02em; }
  header .meta { color: #888; font-size: .9rem; }
  .result { background: #161a23; border: 1px solid #232836; border-radius: 8px;
            padding: 1rem 1.25rem; margin-bottom: 1rem; }
  .result .index { color: #ff5fd7; font-size: .75rem; text-transform: uppercase;
                   letter-spacing: .08em; margin-bottom: .25rem; }
  .result h5 { margin: .1rem 0 .5rem 0; font-size: 1rem; color: #ffffff; }
  .result p  { margin: 0 0 .6rem 0; color: #b8b8b8; line-height: 1.45; }
  .result a  { color: #00d7ff; text-decoration: none; word-break: break-all; }
  .result a:hover { text-decoration: underline; }
</style>
</head>
<body>
  <header>
    <h1>⬡  ScannUs — Search Report</h1>
    <div class="meta">{{ total }} result(s)</div>
  </header>
  {{ results }}
</body>
</html>
"""


class ResultsParser:
    """
    Data transformation layer for search result sets.
    """

    def __init__(self, results: list):
        self.results = results

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _output_path(self, filename: str) -> str:
        """Returns the absolute path inside DIR_REPORTS."""
        return os.path.join(DIR_REPORTS, os.path.basename(filename))

    # ------------------------------------------------------------------
    # HTML export
    # ------------------------------------------------------------------

    def export_html(self, output_path: str) -> None:
        """
        Exports results into a self-contained, styled HTML report.
        """
        path = self._output_path(output_path)

        try:
            fragments = []
            for i, r in enumerate(self.results, 1):
                title = html.escape(r.get("title") or "No title")
                description = html.escape(r.get("description") or "No description")
                link_raw = r.get("link") or ""
                link_attr = html.escape(link_raw, quote=True) or "#"
                link_text = html.escape(link_raw) or "No link"
                fragments.append(
                    '<div class="result">'
                    f'<div class="index">Result {i}</div>'
                    f'<h5>{title}</h5>'
                    f'<p>{description}</p>'
                    f'<a href="{link_attr}" target="_blank" rel="noopener noreferrer">{link_text}</a>'
                    '</div>'
                )

            report = _HTML_TEMPLATE.replace("{{ results }}", "\n".join(fragments))
            report = report.replace("{{ total }}", str(len(self.results)))

            with open(path, "w", encoding="utf-8") as f:
                f.write(report)

            print_success(f"HTML report saved → {path}")

        except IOError as e:
            print_error(f"I/O error exporting HTML: {e}")
        except Exception as e:
            print_error(f"Unexpected error exporting HTML: {e}")

    # ------------------------------------------------------------------
    # CSV export
    # ------------------------------------------------------------------

    def export_csv(self, output_path: str) -> None:
        """Dumps results to a flat UTF-8 CSV file."""
        path = self._output_path(output_path)
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Title", "Description", "Link"])
                for i, r in enumerate(self.results, 1):
                    writer.writerow([
                        i,
                        r.get("title", ""),
                        r.get("description", ""),
                        r.get("link", ""),
                    ])
            print_success(f"CSV saved → {path}")
        except IOError as e:
            print_error(f"I/O error exporting CSV: {e}")
        except Exception as e:
            print_error(f"Unexpected error exporting CSV: {e}")

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_excel(self, output_path: str) -> None:
        """
        Generates a polished Excel .xlsx file with:
        - Styled header row (dark background, white bold text)
        - Alternating row colours (white / light blue)
        - Freeze pane on the first row
        - Auto-fitted column widths
        - Clickable hyperlinks in the Link column
        - Thin borders on all cells
        """
        path = self._output_path(output_path)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            # ── Styles ────────────────────────────────────────────────
            header_fill   = PatternFill("solid", fgColor="1E3A5F")   # dark navy
            header_font   = Font(bold=True, color="FFFFFF", size=11)
            header_align  = Alignment(horizontal="center", vertical="center")

            row_fill_even = PatternFill("solid", fgColor="EAF2FB")   # light blue
            row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")   # white
            row_font      = Font(size=10)
            link_font     = Font(size=10, color="1155CC", underline="single")

            thin = Side(style="thin", color="BFBFBF")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ── Header row ────────────────────────────────────────────
            headers = ["ID", "Title", "Description", "Link"]
            ws.append(headers)
            ws.row_dimensions[1].height = 22

            for col_idx, cell in enumerate(ws[1], 1):
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = border

            # ── Freeze top row ────────────────────────────────────────
            ws.freeze_panes = "A2"

            # ── Data rows ─────────────────────────────────────────────
            for row_num, r in enumerate(self.results, 2):
                fill = row_fill_even if row_num % 2 == 0 else row_fill_odd
                url  = r.get("link", "")

                ws.append([
                    row_num - 1,
                    r.get("title", ""),
                    r.get("description", ""),
                    url,
                ])
                ws.row_dimensions[row_num].height = 16

                for col_idx, cell in enumerate(ws[row_num], 1):
                    cell.fill   = fill
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=(col_idx == 3)  # wrap description column
                    )

                    if col_idx == 4 and url:
                        # Hyperlink in Link column
                        cell.hyperlink = url
                        cell.font = link_font
                    else:
                        cell.font = row_font

            # ── Auto-fit column widths ────────────────────────────────
            col_max_widths = {1: 5, 2: 45, 3: 80, 4: 60}
            for col_idx, col_cells in enumerate(ws.iter_cols(), 1):
                max_len = max(
                    (len(str(cell.value)) if cell.value else 0)
                    for cell in col_cells
                )
                # Cap at defined max to prevent unreasonably wide columns
                fitted = min(max_len + 4, col_max_widths.get(col_idx, 40))
                ws.column_dimensions[get_column_letter(col_idx)].width = fitted

            wb.save(path)
            print_success(f"Excel report saved → {path}")

        except Exception as e:
            print_error(f"Unexpected error exporting Excel: {e}")

    # ------------------------------------------------------------------
    # JSON export
    # ------------------------------------------------------------------

    def export_json(self, output_path: str) -> None:
        """Serialises the result list to a pretty-printed JSON file."""
        path = self._output_path(output_path)
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.results, f, ensure_ascii=False, indent=4)
            print_success(f"JSON saved → {path}")
        except IOError as e:
            print_error(f"I/O error exporting JSON: {e}")
        except Exception as e:
            print_error(f"Unexpected error exporting JSON: {e}")

    # ------------------------------------------------------------------
    # Rich terminal table
    # ------------------------------------------------------------------

    def to_table(self) -> Table:
        """
        Returns a pre-styled Rich Table of the current result set.
        Used for non-interactive rendering in the terminal.
        """
        tbl = make_table(
            f"Results  [{THEME['ACCENT']}]{len(self.results)} found[/]",
            ("#",           THEME["DIM"]),
            ("Title",       "bold white"),
            ("Link",        THEME["LINK"]),
            ("Description", THEME["DIM"]),
            show_lines=True,
        )
        for i, r in enumerate(self.results, 1):
            desc = r.get("description", "")
            tbl.add_row(
                str(i),
                r.get("title", "N/A"),
                r.get("link",  "N/A"),
                (desc[:100] + "…") if len(desc) > 100 else desc,
            )
        return tbl
