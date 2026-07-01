"""
Unit tests for ``utils.results_parse.ResultsParser`` exporters.

Exports normally land in ``core.config.DIR_REPORTS``; each test monkeypatches
that module-level constant to a pytest ``tmp_path`` so nothing is written
outside the sandbox. We then read the files back and assert on structure and,
importantly, on HTML-escaping of untrusted result fields.
"""

import csv
import json

import openpyxl
import pytest

import utils.results_parse as rp
from utils.results_parse import ResultsParser


SAMPLE = [
    {"title": "First", "description": "desc one", "link": "https://a.example/1"},
    {"title": "Second", "description": "desc two", "link": "https://b.example/2"},
]


@pytest.fixture(autouse=True)
def _reports_to_tmp(tmp_path, monkeypatch):
    """Redirect all report output into the test's tmp dir."""
    monkeypatch.setattr(rp, "DIR_REPORTS", str(tmp_path))
    return tmp_path


class TestJSON:
    def test_roundtrip(self, tmp_path):
        ResultsParser(SAMPLE).export_json("out.json")
        data = json.loads((tmp_path / "out.json").read_text(encoding="utf-8"))
        assert data == SAMPLE


class TestCSV:
    def test_header_and_rows(self, tmp_path):
        ResultsParser(SAMPLE).export_csv("out.csv")
        with (tmp_path / "out.csv").open(encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert rows[0] == ["ID", "Title", "Description", "Link"]
        assert rows[1] == ["1", "First", "desc one", "https://a.example/1"]
        assert rows[2][0] == "2"
        assert len(rows) == 3  # header + 2 data rows


class TestHTML:
    def test_contains_content_and_count(self, tmp_path):
        ResultsParser(SAMPLE).export_html("out.html")
        html = (tmp_path / "out.html").read_text(encoding="utf-8")
        assert "First" in html
        assert "https://a.example/1" in html
        assert "2 result(s)" in html

    def test_untrusted_fields_are_escaped(self, tmp_path):
        malicious = [{
            "title": "<script>alert(1)</script>",
            "description": "a & b < c",
            "link": "https://x.example/?q=1&r=2",
        }]
        ResultsParser(malicious).export_html("xss.html")
        html = (tmp_path / "xss.html").read_text(encoding="utf-8")
        # The raw script tag must not appear un-escaped
        assert "<script>alert(1)</script>" not in html
        assert "&lt;script&gt;alert(1)&lt;/script&gt;" in html
        assert "a &amp; b &lt; c" in html


class TestExcel:
    def test_reopens_with_headers_and_rows(self, tmp_path):
        ResultsParser(SAMPLE).export_excel("out.xlsx")
        wb = openpyxl.load_workbook(tmp_path / "out.xlsx")
        ws = wb.active
        assert [c.value for c in ws[1]] == ["ID", "Title", "Description", "Link"]
        # Two data rows after the header
        assert ws.max_row == 3
        assert ws.cell(row=2, column=2).value == "First"
        assert ws.cell(row=3, column=1).value == 2


class TestTable:
    def test_row_and_column_counts(self):
        table = ResultsParser(SAMPLE).to_table()
        assert table.row_count == 2
        assert len(table.columns) == 4

    def test_long_description_truncated(self):
        long_desc = "x" * 250
        table = ResultsParser(
            [{"title": "t", "description": long_desc, "link": "l"}]
        ).to_table()
        # Rendering the last column's cell should be truncated with an ellipsis
        desc_cell = table.columns[3]._cells[0]
        assert "…" in desc_cell
        assert len(desc_cell) < len(long_desc)
